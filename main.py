import pandas as pd
from datetime import datetime, timedelta as time_delta
from random import random, choice
import math
import string
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import *

# Enter your input Excel link here
employees = pd.read_excel(r'/Users/branan/Desktop/ScheduleProject.xlsx', 'Employees').to_dict('records')
Positions = pd.read_excel(r'/Users/branan/Desktop/ScheduleProject.xlsx', 'Positions and Shifts')
TimeOff = pd.read_excel(r'/Users/branan/Desktop/ScheduleProject.xlsx', 'Time off requests').to_dict('records')
Attendance = pd.read_excel(r'/Users/branan/Desktop/ScheduleProject.xlsx', 'Attendance').to_dict('records')
WeeklyInfo = pd.read_excel(r'/Users/branan/Desktop/ScheduleProject.xlsx', 'Weekly Info')

# Manager Parameters

nocallnoshow_weight = 0.6
inadmissiblecallout_weight = 0.2
writeupbehavior_weight = 0.15
late_weight = 0.03
admissiblecallout_weight = 0.02

category1_seniority_weight = 0.7
category1_attendance_weight = 0.3

category5_seniority_weight = 0.3
category5_attendance_weight = 0.7

sigmoid_coefficient = 5



def combine_desired_days(employees):
    days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    for day in days_of_week:
        employees[day + ' Desired'] = (employees[day + ' Desired Position'], employees[day + ' Desired Shift'])
        del employees[day + ' Desired Position']
        del employees[day + ' Desired Shift']
    return employees


def separate_positions_eligible(positions_string):
    positions = positions_string.split(',')
    positions = [position.strip() for position in positions]
    return tuple(positions)


def days_from_today(timestamp):
    timestamp = str(timestamp)
    date = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
    today = datetime.now()
    difference = today - date
    return difference.days


# Inputs any format of shift and returns a tuple with start, end, and hours
def timedelta(shift_hours: str) -> tuple:
    start_time, end_time = shift_hours.split('-')
    start_time = start_time.strip()
    end_time = end_time.strip()

    formats = ['%I:%M %p', '%I:%M%p', '%I%p', '%I %p', '%I:%M', '%I', '%H:%M %p', '%H:%M%p', '%H%p', '%H %p', '%H:%M',
               '%H']
    for fmt in formats:
        try:
            start_datetime = datetime.strptime(start_time, fmt)
            break
        except ValueError:
            continue
    for fmt in formats:
        try:
            end_datetime = datetime.strptime(end_time, fmt)
            break
        except ValueError:
            continue

    start_time = start_datetime.time()
    end_time = end_datetime.time()
    shift_duration = end_datetime - start_datetime
    start_time_str = start_datetime.strftime('%I:%M %p')
    end_time_str = end_datetime.strftime('%I:%M %p')
    duration_hours = shift_duration.seconds // 3600
    duration_minutes = (shift_duration.seconds // 60) % 60
    duration = duration_hours + duration_minutes / 60
    return start_time_str, end_time_str, duration


# Returns true if time1 is earlier than time2
def earlier_than(time1: str, time2: str) -> bool:
    formats = ['%I:%M %p', '%I:%M%p', '%I%p', '%I %p', '%I:%M', '%I', '%H:%M %p', '%H:%M%p', '%H%p', '%H %p', '%H:%M',
               '%H']
    for fmt in formats:
        try:
            time1_datetime = datetime.strptime(time1, fmt)
            break
        except ValueError:
            continue
    for fmt in formats:
        try:
            time2_datetime = datetime.strptime(time2, fmt)
            break
        except ValueError:
            continue
    return time1_datetime < time2_datetime


# Returns true if shift 1 is within shift 2
def within(shift1: str, shift2: str) -> bool:
    start1, end1, duration1 = timedelta(shift1)
    start2, end2, duration2 = timedelta(shift2)
    if earlier_than(start1, start2) or earlier_than(end2, start1):
        return False
    if earlier_than(end1, start2) or earlier_than(end2, end1):
        return False
    return True


def same_hours(shift1: str, shift2: str) -> bool:
    start1, end1, duration1 = timedelta(shift1)
    start2, end2, duration2 = timedelta(shift2)
    if earlier_than(start1, start2) or earlier_than(end2, start1):
        return False
    if earlier_than(end1, start2) or earlier_than(end2, end1):
        return False
    if start1 == start2 and end1 == end2:
        return True
    return False


# Same as within but accounts for off days
def is_available(shift1: str, shift2: str) -> bool:
    if shift1 == 'open' or shift1 == 'Open':
        return True
    elif shift1 == 'off' or shift1 == 'Off':
        return False
    else:
        return within(shift2, shift1)


# Returns true if employee position AND particular shift is eligible.
def is_eligible(employee_name, shift_info):
    for employee in employees:
        if employee['Name'] == employee_name:
            position_eligible = shift_info[0] in employee['Positions Eligible']
            for eligible_shift in employee['Shifts Eligible']:
                if eligible_shift == 'All' or eligible_shift == 'all':
                    shift_eligible = True
                    break
                else:
                    shift_eligible = same_hours(shift_info[1], eligible_shift)
                    break

    return position_eligible and shift_eligible


def days_between(date1, date2):
    date1 = datetime.strptime(str(date1), '%Y-%m-%d %H:%M:%S')
    date2 = datetime.strptime(str(date2), '%Y-%m-%d %H:%M:%S')
    difference = date2 - date1
    return difference.days + 1


def is_between(start_date_str, end_date_str, test_date_str):
    start_date = datetime.strptime(start_date_str, '%m/%d/%Y')
    end_date = datetime.strptime(end_date_str, '%m/%d/%Y')
    test_date = datetime.strptime(test_date_str, '%m/%d/%Y')

    return start_date <= test_date <= end_date


def count_noncompliances_in_time_range(employee, noncompliance, days_since, time_range):
    # Check if the noncompliance occurred within the specified time range
    if days_since <= time_range:
        # Increment the count for the noncompliance in the specified time range
        if f"{noncompliance} in past {time_range} days" in employee:
            employee[f"{noncompliance} in past {time_range} days"] += 1
        else:
            employee[f"{noncompliance} in past {time_range} days"] = 1


# Refreshes 'Number of shifts eligible' and 'Hours eligible'. Employees are eligible for a shift and hours if they
#       are not already assigned to that day of the week, and that shift is 'Unassigned'
def update_assignments():
    unassigned_count = 0
    for employee in employees:
        employee['Number of shifts assigned'] = 0
        employee['Hours assigned'] = 0
        employee['Number of shifts eligible'] = 0
        employee['Hours eligible'] = 0

    # Add the number of shifts and hours eligible for each employee
    for name, shifts in Employee_eligibility.items():
        employee = next(emp for emp in employees if emp['Name'] == name)
        employee['Number of shifts eligible'] = len(
            [shift for shift in shifts if shift[1] in [s[1] for s in Assignments if s[4] == 'Unassigned']])
        employee['Hours eligible'] = sum(
            [shift[3] for shift in shifts if shift[1] in [s[1] for s in Assignments if s[4] == 'Unassigned']])

    # Subtract the number of shifts and hours assigned from the number of shifts and hours eligible for each employee
    for shift in Assignments:
        if shift[4] != 'Unassigned':
            employee = next(emp for emp in employees if emp['Name'] == shift[4])
            employee['Number of shifts assigned'] += 1
            employee['Hours assigned'] += shift[3]
            day_of_week = shift[1].split()[0]
            employee['Number of shifts eligible'] -= len(
                [shift for shift in Employee_eligibility[employee['Name']] if shift[1].startswith(day_of_week)]) - 1
            employee['Hours eligible'] -= sum(
                [shift[3] for shift in Employee_eligibility[employee['Name']] if shift[1].startswith(day_of_week)]) - \
                                          shift[3]
    for employee in employees:
        for assign in Assignments:
            if assign[4] == employee['Name']:
                i = 0
                while i < len(Employee_eligibility[employee['Name']]):
                    shift = Employee_eligibility[employee['Name']][i]
                    if shift[1].split()[0] == assign[1].split()[0]:
                        del Employee_eligibility[employee['Name']][i]
                    else:
                        i += 1
    for shift in Assignments:
        if shift[4] == 'Unassigned':
            unassigned_count += 1
    return unassigned_count


def sigmoid(x):
    return 1 / (1 + math.exp(-x))


def max_score(lst):
    max_name = lst[0][0]
    max_val = lst[0][1]
    second_max_val = 0
    for name, val in lst[1:]:
        if val > max_val:
            max_name = name
            second_max_val = max_val
            max_val = val
        elif val > second_max_val:
            second_max_val = val
    difference = max_val - second_max_val
    if difference < 0.05:
        probability = 0.5
    else:
        probability = sigmoid(sigmoid_coefficient*difference)
    names_with_scores_other_than_max = [name for name, val in lst if val != max_val]
    if random() < probability and names_with_scores_other_than_max:
        return choice(names_with_scores_other_than_max)
    else:
        return max_name


def min_score(lst):
    min_name = lst[0][0]
    min_val = lst[0][1]
    second_min_val = 1
    for name, val in lst[1:]:
        if val < min_val:
            min_name = name
            second_min_val = min_val
            min_val = val
        elif val < second_min_val:
            second_min_val = val
    difference = second_min_val - min_val
    if difference < 0.05:
        probability = 0.5
    else:
        probability = sigmoid((difference - 0.05) / 0.2)
    names_with_scores_other_than_min = [name for name, val in lst if val != min_val]
    if random() < probability and names_with_scores_other_than_min:
        return choice(names_with_scores_other_than_min)
    else:
        return min_name


def no_request_off(date_of_shift, name):
    shift_found = True
    type = 'none'
    for request in TimeOff:
        if request['Name'] == name:
            if is_between(request['First day requested'], request['Last day requested'], date_of_shift):
                shift_found = False
                type = request['Type']
    return shift_found, type


def hyphenate(s):
    if len(s) != 10:
        raise ValueError("Input string must be exactly 10 characters long")
    return s[:2] + "-" + s[3:5] + "-" + s[6:]


def reformat_date(date_string: str) -> str:
    # Parse the date string into a datetime object
    date = datetime.strptime(date_string, "%m/%d/%Y")

    # Format the date using the "%b %d" format
    day = date.strftime("%d").lstrip("0")
    formatted_date = f"{date.strftime('%b')} {day}"

    return formatted_date


def alphabetize_by_last_name(names):
    # Sort the list of names using a lambda function as the key
    names.sort(key=lambda x: x.split()[1])
    return names


# Organizing the Employee data

# Combining Desired Shift and Position into a tuple
employees = [combine_desired_days(employee) for employee in employees]

for employee in employees:
    # Changing Positions eligible for a tuple
    positions_string = employee['Position(s) Eligible']
    employee['Positions Eligible'] = separate_positions_eligible(positions_string)
    del employee['Position(s) Eligible']
    # Changing Date Hired to days since hired
    employee['Days Since Hire'] = days_from_today(employee['Date Hired'])
    del employee['Date Hired']

for employee in employees:
    shifts_string = employee['Shifts Eligible']
    if shifts_string == 'All':
        employee['Shifts Eligible'] = ('All',)
    else:
        shifts_list = shifts_string.split(',')
        shifts_tuple = tuple(shift.strip() for shift in shifts_list)
        employee['Shifts Eligible'] = shifts_tuple

# Organizing the Positions and Shifts data
positions_dict = {}
for i in range(len(Positions)):
    if i % 10 == 0:
        position_name = Positions.iloc[i, 0]
        if pd.notnull(position_name):
            positions_dict[position_name] = {}
            positions_dict[position_name]['Position Name'] = position_name
            shift_count = 0
            for j in range(2, len(Positions.columns)):
                if pd.notnull(Positions.iloc[i, j]):
                    shift_count += 1
            positions_dict[position_name]['Number of Shifts'] = shift_count
    else:
        for j in range(2, len(Positions.columns)):
            if pd.notnull(Positions.iloc[i, j]):
                positions_dict[position_name][f'Shift {j - 1} Hours'] = Positions.iloc[i - (i % 10) + 1, j]

days_of_week = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
for position, i in zip(positions_dict.keys(), range(0, len(positions_dict.keys()))):
    for shift in range(1, (positions_dict[position]['Number of Shifts'] + 1)):
        for day, index in zip(days_of_week, range(2, 9)):
            key = f"{day} Shift {shift} Employees"
            positions_dict[position][key] = Positions.iloc[(index + 10 * i), (shift + 1)]

# Organizing Attendance Data
for employee in Attendance:
    employee['Days Since Infraction'] = days_from_today(employee['Date'])
    del employee['Date']

for employee in employees:
    employee['Admissible call-out'] = 0
    employee['Inadmissible call-out'] = 0
    employee['Late'] = 0
    employee['No call no show'] = 0
    employee['Write-up behavior'] = 0
    employee['Admissible call-out in past 30 days'] = 0
    employee['Inadmissible call-out in past 30 days'] = 0
    employee['Late in past 30 days'] = 0
    employee['No call no show in past 30 days'] = 0
    employee['Write-up behavior in past 30 days'] = 0
    employee['Admissible call-out in past 90 days'] = 0
    employee['Inadmissible call-out in past 90 days'] = 0
    employee['Late in past 90 days'] = 0
    employee['No call no show in past 90 days'] = 0
    employee['Write-up behavior in past 90 days'] = 0

for employee in Attendance:
    found = False
    for e in employees:
        if e['Name'] == employee['Name']:
            # Increment the overall count for the noncompliance
            if employee['Noncompliance'] in e:
                e[employee['Noncompliance']] += 1
            else:
                e[employee['Noncompliance']] = 1
            # Count the number of noncompliances in the past 30 and 90 days
            count_noncompliances_in_time_range(e, employee['Noncompliance'], employee['Days Since Infraction'], 30)
            count_noncompliances_in_time_range(e, employee['Noncompliance'], employee['Days Since Infraction'], 90)
            found = True
            break
    if not found:
        employee_entry = {'Name': employee['Name'], employee['Noncompliance']: 1}
        count_noncompliances_in_time_range(employee_entry, employee['Noncompliance'], employee['Days Since Infraction'],
                                           30)
        count_noncompliances_in_time_range(employee_entry, employee['Noncompliance'], employee['Days Since Infraction'],
                                           90)
        employees.append(employee_entry)

# Organizing Time off Data
for request in TimeOff:
    request['Days Requested'] = days_between(request['First day requested'], request['Last day requested'])
    request['First day requested'] = request['First day requested'].strftime('%m/%d/%Y')
    request['Last day requested'] = request['Last day requested'].strftime('%m/%d/%Y')

for employee in employees:
    employee['Approved Days Off in past 3 months'] = 0
    employee['Approved Days Off in past 6 months'] = 0
    employee['Rejected Time Off Requests in past 3 months'] = 0
    employee['Rejected Time Off Requests in past 6 months'] = 0
    employee['Approved Days Off Lifetime'] = 0
    employee['Rejected Time Off Requests Lifetime'] = 0

for request in TimeOff:
    if request['Type'] == 'Approved':
        days = request['Days Requested']
        for employee in employees:
            if employee['Name'] == request['Name']:
                employee['Approved Days Off Lifetime'] += days
                break
    if request['Type'] == 'Rejected':
        for employee in employees:
            if employee['Name'] == request['Name']:
                employee['Rejected Time Off Requests Lifetime'] += 1
                break

for request in TimeOff:
    date = datetime.strptime(request['First day requested'], '%m/%d/%Y')
    today = datetime.now()
    difference = today - date
    days = difference.days
    if days <= 90:
        for employee in employees:
            if employee['Name'] == request['Name']:
                if request['Type'] == 'Approved':
                    employee['Approved Days Off in past 3 months'] += request['Days Requested']
                if request['Type'] == 'Rejected':
                    employee['Rejected Time Off Requests in past 3 months'] += 1
    if days <= 180:
        for employee in employees:
            if employee['Name'] == request['Name']:
                if request['Type'] == 'Approved':
                    employee['Approved Days Off in past 6 months'] += request['Days Requested']
                if request['Type'] == 'Rejected':
                    employee['Rejected Time Off Requests in past 6 months'] += 1

# Organizing date to weekday data
MondayDate = WeeklyInfo.iloc[0, 1]
MondayDate = MondayDate.strftime('%m/%d/%Y')

start_date = datetime.strptime(MondayDate, '%m/%d/%Y')
day_names = {'Monday': 0, 'Tuesday': 1, 'Wednesday': 2, 'Thursday': 3, 'Friday': 4, 'Saturday': 5, 'Sunday': 6}

for day_name in day_names:
    day_index = day_names[day_name]
    day_date = start_date + time_delta(days=day_index)
    day_date_str = day_date.strftime('%m/%d/%Y')
    globals()[f'{day_name}Date'] = day_date_str

# Creates list of Assignments with 'Unassigned'
# Creates Employee_eligibility with employee lists containing only shifts if is_eligible AND is_available is true
Assignments = []
Employee_eligibility = {}

for position, shifts in positions_dict.items():
    position_name = shifts['Position Name']
    num_shifts = shifts['Number of Shifts']
    for shift, num_employees in shifts.items():
        if 'Shift' in shift and 'Employees' in shift and 1 <= int(shift.split()[-2]) <= num_shifts:
            shift_number = int(shift.split()[-2])
            shift_hours = shifts[f'Shift {shift_number} Hours']
            duration = timedelta(shift_hours)[2]
            for i in range(num_employees):
                Assignments.append((position_name, shift, shift_hours, duration, 'Unassigned'))

for i, shift in enumerate(Assignments):
    position = shift[0]
    shift_hours = shift[2]
    day_of_week = shift[1].split()[0]
    shift_number = shift[1].split()[2]
    availability_key = f"{day_of_week} Availability"
    for employee in employees:
        employee['Number of shifts eligible'] = 0
        employee['Hours eligible'] = 0
        name = employee['Name']
        availability = employee[availability_key]
        if is_eligible(name, (position, shift_hours)) and is_available(availability, shift_hours) and shift[
            4] == 'Unassigned':
            if name not in Employee_eligibility:
                Employee_eligibility[name] = []
            Employee_eligibility[name].append((shift[0], f"{day_of_week} Shift {shift_number}", shift[2], shift[3]))
            new_shift = (shift[0], f"{day_of_week} Shift {shift_number}", shift[2], shift[3], shift[4])
            Assignments[i] = new_shift

# Attendance score
for employee in employees:
    employee['Attendance score'] = 1 - (
            0 * employee['No call no show in past 30 days'] +
            0 * employee['No call no show in past 90 days'] +
            nocallnoshow_weight * employee['No call no show'] +
            0 * employee['Inadmissible call-out in past 30 days'] +
            0 * employee['Inadmissible call-out in past 90 days'] +
            inadmissiblecallout_weight * employee['Inadmissible call-out'] +
            0 * employee['Write-up behavior in past 30 days'] +
            0 * employee['Write-up behavior in past 90 days'] +
            writeupbehavior_weight * employee['Write-up behavior'] +
            0 * employee['Late in past 30 days'] +
            0 * employee['Late in past 90 days'] +
            late_weight * employee['Late'] +
            0 * employee['Admissible call-out in past 30 days'] +
            0 * employee['Admissible call-out in past 90 days'] +
            admissiblecallout_weight * employee['Admissible call-out'])

# THE ALGORITHM

# Assigns employees if they are eligible, the only one who prefers a shift in Assignments, and assigning that shift won't put them over their max hours/shifts
update_assignments()
while update_assignments() > 0:
    assignments_copy = Assignments.copy()
    for assignment in assignments_copy:
        preferred_list = []
        if assignment[4] == "Unassigned":
            # Checking if eligible
            for name, shifts in Employee_eligibility.items():
                for shift in shifts:
                    position, shift_name, shift_hours, shift_duration = shift
                    if assignment[1] == shift_name and assignment[0] == position:
                        day_of_week = shift_name.split()[0]
                        date_of_shift = locals()[day_of_week + 'Date']
                        for employee in employees:
                            if employee['Name'] == name:
                                preferred_shift = employee[f'{day_of_week} Desired']
                                # Checking if requested off
                                if no_request_off(date_of_shift, name)[0]:
                                    # Checking if preferred and added shift doesn't exceed max hours
                                    if preferred_shift[0] == assignment[0] and same_hours(preferred_shift[1],
                                                                                          assignment[2]) and employee[
                                        'Hours assigned'] < (employee['Max Desired Hours'] - shift_duration) and \
                                            employee['Number of shifts assigned'] < (employee['Max # Days'] - 1):
                                        preferred_list.append(name)
            if len(preferred_list) == 1:
                for employee in employees:
                    if employee['Name'] == preferred_list[0]:
                        try:
                            index = Assignments.index(assignment)
                        except ValueError:
                            continue
                        Assignments[index] = (
                            assignment[0], assignment[1], assignment[2], assignment[3], preferred_list[0])
                        update_assignments()

    # For shifts that have multiple eligible prefers, assigns employee with max score (seniority and attendance)
    assignments_copy = Assignments.copy()
    for assignment in assignments_copy:
        preferred_list = []
        if assignment[4] == "Unassigned":
            # Checking if eligible
            for name, shifts in Employee_eligibility.items():
                for shift in shifts:
                    position, shift_name, shift_hours, shift_duration = shift
                    if assignment[1] == shift_name and assignment[0] == position:
                        day_of_week = shift_name.split()[0]
                        date_of_shift = locals()[day_of_week + 'Date']
                        for employee in employees:
                            if employee['Name'] == name:
                                preferred_shift = employee[f'{day_of_week} Desired']
                                # Checking if requested off
                                if no_request_off(date_of_shift, name)[0]:
                                    # Checking if preferred and added shift doesn't exceed max hours
                                    if preferred_shift[0] == assignment[0] and same_hours(preferred_shift[1],assignment[2]) and employee['Hours assigned'] < (employee['Max Desired Hours'] - shift_duration) and employee['Number of shifts assigned'] < (employee['Max # Days'] - 1):
                                        preferred_list.append(name)
            if len(preferred_list) > 1:
                comparison_list = []
                longest_hire = 0
                max_attendance_score = 0
                for employee in employees:
                    if employee['Name'] in preferred_list:
                        if employee['Days Since Hire'] > longest_hire:
                            longest_hire = employee['Days Since Hire']
                        if employee['Attendance score'] > max_attendance_score:
                            max_attendance_score = employee['Attendance score']
                for employee in employees:
                    if employee['Name'] in preferred_list:
                        comparison_score = category1_seniority_weight * (employee['Days Since Hire'] / longest_hire) + category1_attendance_weight * (
                                employee['Attendance score'] / max_attendance_score)
                        comparison_list.append((employee['Name'], comparison_score))
                winner = max_score(comparison_list)
                # Assigning winner
                try:
                    index = Assignments.index(assignment)
                except ValueError:
                    continue
                Assignments[index] = (assignment[0], assignment[1], assignment[2], assignment[3], winner)
                update_assignments()

    # Eligible employees who don't prefer OFF, haven't requested time off, added shift !> hours/shifts
    assignments_copy = Assignments.copy()
    for assignment in assignments_copy:
        preferred_list = []
        if assignment[4] == "Unassigned":
            # Checking if eligible
            for name, shifts in Employee_eligibility.items():
                for shift in shifts:
                    position, shift_name, shift_hours, shift_duration = shift
                    if assignment[1] == shift_name and assignment[0] == position:
                        day_of_week = shift_name.split()[0]
                        date_of_shift = locals()[day_of_week + 'Date']
                        for employee in employees:
                            if employee['Name'] == name:
                                preferred_shift = employee[f'{day_of_week} Desired']
                                # Checking if requested off
                                if no_request_off(date_of_shift, name)[0]:
                                    # Checking if not preferred off and added shift doesn't exceed max hours
                                    if preferred_shift[0].upper() != "OFF" and employee['Hours assigned'] < (
                                            employee['Max Desired Hours'] - shift_duration) and employee['Number of shifts assigned'] < (employee['Max # Days'] - 1):
                                        preferred_list.append(name)
            if len(preferred_list) == 1:
                for employee in employees:
                    if employee['Name'] == preferred_list[0]:
                        try:
                            index = Assignments.index(assignment)
                        except ValueError:
                            continue
                        Assignments[index] = (
                            assignment[0], assignment[1], assignment[2], assignment[3], preferred_list[0])
                        update_assignments()
            if len(preferred_list) > 1:
                if len(preferred_list) > 1:
                    comparison_list = []
                    largest_proportion = 0.01
                    for employee in employees:
                        if employee['Name'] in preferred_list:
                            if employee['Min Desired Hours'] != 0:
                                if employee['Hours assigned'] / employee['Min Desired Hours'] > largest_proportion:
                                    largest_proportion = employee['Hours assigned'] / employee['Min Desired Hours']
                    for employee in employees:
                        if employee['Name'] in preferred_list:
                            if employee['Min Desired Hours'] == 0:
                                comparison_score = 0.01
                            else:
                                comparison_score = (employee['Hours assigned'] / employee[
                                    'Min Desired Hours']) / largest_proportion
                            comparison_list.append((employee['Name'], comparison_score))
                    winner = min_score(comparison_list)
                    # Assigning winner
                    try:
                        index = Assignments.index(assignment)
                    except ValueError:
                        continue
                    Assignments[index] = (assignment[0], assignment[1], assignment[2], assignment[3], winner)
                    update_assignments()

    # Assigns employees who prefer 'OFF' but are eligible, haven't requested off, and added shift !> max hours
    assignments_copy = Assignments.copy()
    for assignment in assignments_copy:
        preferred_list = []
        if assignment[4] == "Unassigned":
            # Checking if eligible
            for name, shifts in Employee_eligibility.items():
                for shift in shifts:
                    position, shift_name, shift_hours, shift_duration = shift
                    if assignment[1] == shift_name and assignment[0] == position:
                        day_of_week = shift_name.split()[0]
                        date_of_shift = locals()[day_of_week + 'Date']
                        for employee in employees:
                            if employee['Name'] == name:
                                preferred_shift = employee[f'{day_of_week} Desired']
                                # Checking if requested off
                                if no_request_off(date_of_shift, name)[0]:
                                    # Checking if added shift doesn't exceed max hours or 1 more than max days
                                    if employee['Hours assigned'] < (employee['Max Desired Hours'] - shift_duration) and employee['Number of shifts assigned'] < (employee['Max # Days'] - 1):
                                        preferred_list.append(name)
            if len(preferred_list) == 1:
                for employee in employees:
                    if employee['Name'] == preferred_list[0]:
                        try:
                            index = Assignments.index(assignment)
                        except ValueError:
                            continue
                        Assignments[index] = (
                            assignment[0], assignment[1], assignment[2], assignment[3], preferred_list[0])
                        update_assignments()
            if len(preferred_list) > 1:
                if len(preferred_list) > 1:
                    comparison_list = []
                    largest_proportion = 0.01
                    for employee in employees:
                        if employee['Name'] in preferred_list:
                            if employee['Min Desired Hours'] != 0:
                                if employee['Hours assigned'] / employee['Min Desired Hours'] > largest_proportion:
                                    largest_proportion = employee['Hours assigned'] / employee['Min Desired Hours']
                    for employee in employees:
                        if employee['Name'] in preferred_list:
                            if employee['Min Desired Hours'] == 0:
                                comparison_score = 0
                            else:
                                comparison_score = (employee['Hours assigned'] / employee['Min Desired Hours']) / largest_proportion
                            comparison_list.append((employee['Name'], comparison_score))
                    winner = min_score(comparison_list)
                    # Assigning winner
                    try:
                        index = Assignments.index(assignment)
                    except ValueError:
                        continue
                    Assignments[index] = (assignment[0], assignment[1], assignment[2], assignment[3], winner)
                    update_assignments()

    # Assigns employees that would go no more than 5 hours over their max. If multiple assigns based on 33% seniority, 33% avg hours, 33% attendance, minimum score.
    assignments_copy = Assignments.copy()
    for assignment in assignments_copy:
        preferred_list = []
        if assignment[4] == "Unassigned":
            # Checking if eligible
            for name, shifts in Employee_eligibility.items():
                for shift in shifts:
                    position, shift_name, shift_hours, shift_duration = shift
                    if assignment[1] == shift_name and assignment[0] == position:
                        day_of_week = shift_name.split()[0]
                        date_of_shift = locals()[day_of_week + 'Date']
                        for employee in employees:
                            if employee['Name'] == name:
                                if no_request_off(date_of_shift, name)[0] == True:
                                    if employee['Hours assigned'] < (employee['Max Desired Hours'] - shift_duration + 5) and employee['Number of shifts assigned'] < employee['Max # Days']:
                                        preferred_list.append(name)
            if len(preferred_list) == 1:
                for employee in employees:
                    if employee['Name'] == preferred_list[0]:
                        try:
                            index = Assignments.index(assignment)
                        except ValueError:
                            continue
                        Assignments[index] = (
                            assignment[0], assignment[1], assignment[2], assignment[3], preferred_list[0])
                        update_assignments()
            if len(preferred_list) > 1:
                comparison_list = []
                largest_proportion = 0.01
                for employee in employees:
                    if employee['Name'] in preferred_list:
                        if employee['Min Desired Hours'] != 0:
                            if employee['Hours assigned'] / employee['Min Desired Hours'] > largest_proportion:
                                largest_proportion = employee['Hours assigned'] / employee['Min Desired Hours']
                for employee in employees:
                    if employee['Name'] in preferred_list:
                        if employee['Min Desired Hours'] == 0:
                            comparison_score = 0.01
                        else:
                            comparison_score = (employee['Hours assigned'] / employee[
                                'Min Desired Hours']) / largest_proportion
                        comparison_list.append((employee['Name'], comparison_score))
                winner = min_score(comparison_list)
                # Assigning winner
                try:
                    index = Assignments.index(assignment)
                except ValueError:
                    continue
                Assignments[index] = (assignment[0], assignment[1], assignment[2], assignment[3], winner)
                update_assignments()

    # Assigns employees that have if possible request off
    assignments_copy = Assignments.copy()
    for assignment in assignments_copy:
        preferred_list = []
        if assignment[4] == "Unassigned":
            # Checking if eligible
            for name, shifts in Employee_eligibility.items():
                for shift in shifts:
                    position, shift_name, shift_hours, shift_duration = shift
                    if assignment[1] == shift_name and assignment[0] == position:
                        day_of_week = shift_name.split()[0]
                        date_of_shift = locals()[day_of_week + 'Date']
                        for employee in employees:
                            if employee['Name'] == name:
                                if no_request_off(date_of_shift, name) == (False, 'If possible') or \
                                        no_request_off(date_of_shift, name)[0] == True:
                                    if employee['Hours assigned'] < (employee['Max Desired Hours'] - shift_duration + 5) and employee['Number of shifts assigned'] < employee['Max # Days']:
                                        preferred_list.append(name)
            if len(preferred_list) == 1:
                for employee in employees:
                    if employee['Name'] == preferred_list[0]:
                        try:
                            index = Assignments.index(assignment)
                        except ValueError:
                            continue
                        Assignments[index] = (
                            assignment[0], assignment[1], assignment[2], assignment[3], preferred_list[0])
                        update_assignments()
            if len(preferred_list) > 1:
                comparison_list = []
                longest_hire = 0
                max_attendance_score = 0
                for employee in employees:
                    if employee['Name'] in preferred_list:
                        if employee['Days Since Hire'] > longest_hire:
                            longest_hire = employee['Days Since Hire']
                        if employee['Attendance score'] > max_attendance_score:
                            max_attendance_score = employee['Attendance score']
                for employee in employees:
                    if employee['Name'] in preferred_list:
                        comparison_score = category5_seniority_weight * (employee['Days Since Hire'] / longest_hire) + category5_attendance_weight * (
                                employee['Attendance score'] / max_attendance_score)
                        comparison_list.append((employee['Name'], comparison_score))
                winner = min_score(comparison_list)
                # Assigning winner
                try:
                    index = Assignments.index(assignment)
                except ValueError:
                    continue
                Assignments[index] = (assignment[0], assignment[1], assignment[2], assignment[3], winner)
                update_assignments()
    break


# Preparing export
positions_list = []
for position in positions_dict:
    positions_list.append(position)

for i, position in enumerate(positions_list):
    num_position = 0
    name_list = []
    for employee in employees:
        if position in employee['Positions Eligible']:
            num_position += 1
            name_list.append(employee['Name'])
    positions_list[i] = [(position, num_position), alphabetize_by_last_name(name_list)]

MondayDate2 = hyphenate(MondayDate)
SundayDate2 = hyphenate(SundayDate)
CompanyName = WeeklyInfo.iloc[1, 1]


# Export excel file
workbook = Workbook()
sheet = workbook.active
sheet.sheet_view.zoomScale = 150
sheet.row_dimensions[1].height = 34
sheet.column_dimensions['A'].width = 18

for column in string.ascii_uppercase[1:8]:
    sheet.column_dimensions[column].width = 15.83

sheet.merge_cells('A1:H1')
sheet["A1"] = CompanyName
sheet["A1"].alignment = Alignment(horizontal='center', vertical='center')
sheet['A1'].font = Font(size=26, bold=True)
sheet["A1"].fill = PatternFill("solid", start_color="79afd4")

sheet["B2"] = reformat_date(MondayDate)
sheet["C2"] = reformat_date(TuesdayDate)
sheet["D2"] = reformat_date(WednesdayDate)
sheet["E2"] = reformat_date(ThursdayDate)
sheet["F2"] = reformat_date(FridayDate)
sheet["G2"] = reformat_date(SaturdayDate)
sheet["H2"] = reformat_date(SundayDate)

sheet["B3"] = "Monday"
sheet["C3"] = "Tuesday"
sheet["D3"] = "Wednesday"
sheet["E3"] = "Thursday"
sheet["F3"] = "Friday"
sheet["G3"] = "Saturday"
sheet["H3"] = "Sunday"

for row in sheet['B2:H3']:
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

for row in sheet['B2:H2']:
    for cell in row:
        cell.fill = PatternFill("solid", start_color="e3cd8f")

# for position_number in range(0, len(positions_list)):

starting_row = 4
for i in range(0, len(positions_list)):
    sheet.merge_cells(f'A{starting_row}:H{starting_row}')
    sheet[f"A{starting_row}"] = positions_list[i][0][0]
    sheet[f"A{starting_row}"].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f"A{starting_row}"].fill = PatternFill("solid", start_color="79afd4")

    for employee, rows in zip(positions_list[i][1],
                              range(starting_row + 1, starting_row + 1 + positions_list[i][0][1])):
        assignments_list = []
        for day in days_of_week:
            found = False
            for shift in Assignments:
                if shift[4] == employee and shift[1].split()[0] == day and shift[0] == positions_list[i][0][0]:
                    assignments_list.append(shift[2])
                    found = True
                    break
            if not found:
                toggle = False
                while not toggle:
                    if day == 'Monday':
                        if no_request_off(MondayDate, employee) == (False, 'Approved') or no_request_off(MondayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    elif day == 'Tuesday':
                        if no_request_off(TuesdayDate, employee) == (False, 'Approved') or no_request_off(TuesdayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    elif day == 'Wednesday':
                        if no_request_off(WednesdayDate, employee) == (False, 'Approved') or no_request_off(WednesdayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    elif day == 'Thursday':
                        if no_request_off(ThursdayDate, employee) == (False, 'Approved') or no_request_off(ThursdayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    elif day == 'Friday':
                        if no_request_off(FridayDate, employee) == (False, 'Approved') or no_request_off(FridayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    elif day == 'Saturday':
                        if no_request_off(SaturdayDate, employee) == (False, 'Approved') or no_request_off(SaturdayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    elif day == 'Sunday':
                        if no_request_off(SundayDate, employee) == (False, 'Approved') or no_request_off(SundayDate, employee) == (False, 'If possible'):
                            assignments_list.append("PTO")
                            toggle = True
                    assignments_list.append("OFF")
                    toggle = True

        for shift, letter in zip(assignments_list, ['B', 'C', 'D', 'E', 'F', 'G', 'H']):
            inner_cell = f"{letter}{rows}"
            sheet[inner_cell] = shift
            sheet[inner_cell].alignment = Alignment(horizontal='center', vertical='center')

            if shift == 'OFF':
                sheet[inner_cell].fill = PatternFill("solid", start_color="f0f56e")
            if shift == 'PTO':
                sheet[inner_cell].fill = PatternFill("solid", start_color="f27c41")
        cell = f"A{rows}"
        sheet[cell] = employee

    starting_row = starting_row + 1 + positions_list[i][0][1]

for row in sheet[f'A1:H{starting_row - 1}']:
    for cell in row:
        cell.border = openpyxl.styles.borders.Border(
            top=openpyxl.styles.borders.Side(style='thin'),
            left=openpyxl.styles.borders.Side(style='thin'),
            right=openpyxl.styles.borders.Side(style='thin'),
            bottom=openpyxl.styles.borders.Side(style='thin'))

unassigned_count = 0
for assignment in Assignments:
    if assignment[4] == 'Unassigned':
        unassigned_count += 1

if unassigned_count > 0:

    sheet[f'A{starting_row + 1}'] = "Unassigned Shifts:"

    row_b = starting_row + 2
    row_c = starting_row + 2
    row_d = starting_row + 2
    row_e = starting_row + 2
    row_f = starting_row + 2
    row_g = starting_row + 2
    row_h = starting_row + 2

    for assignment in Assignments:
        if assignment[4] == 'Unassigned':
            for day, letter in zip(days_of_week, ['B', 'C', 'D', 'E', 'F', 'G', 'H']):
                if assignment[1].split()[0] == day:
                    if letter == 'B':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_b}'] = cell_text
                        sheet[f'{letter}{row_b}'].alignment = Alignment(wrap_text=True)
                        row_b += 1
                    elif letter == 'C':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_c}'] = cell_text
                        sheet[f'{letter}{row_c}'].alignment = Alignment(wrap_text=True)
                        row_c += 1
                    elif letter == 'D':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_d}'] = cell_text
                        sheet[f'{letter}{row_d}'].alignment = Alignment(wrap_text=True)
                        row_d += 1
                    elif letter == 'E':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_e}'] = cell_text
                        sheet[f'{letter}{row_e}'].alignment = Alignment(wrap_text=True)
                        row_e += 1
                    elif letter == 'F':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_f}'] = cell_text
                        sheet[f'{letter}{row_f}'].alignment = Alignment(wrap_text=True)
                        row_f += 1
                    elif letter == 'G':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_g}'] = cell_text
                        sheet[f'{letter}{row_g}'].alignment = Alignment(wrap_text=True)
                        row_g += 1
                    elif letter == 'H':
                        text_list = f"{assignment[0]},{assignment[2]}".split(',')
                        text_list[1] = '\n' + text_list[1]
                        cell_text = ','.join(text_list)

                        sheet[f'{letter}{row_h}'] = cell_text
                        sheet[f'{letter}{row_h}'].alignment = Alignment(wrap_text=True)
                        row_h += 1

workbook.save("Export.xlsx")
print(openpyxl.__version__)
